from openpyxl import Workbook
wb = Workbook()
# 활성화된 시트
ws = wb.active
ws.title = 'BK_'
# 새로운 시트 (기본 이름 생성)
ws = wb.create_sheet()
ws.title = 'BK_new'
# 시트 탭 컬러
ws.sheet_properties.tabColor = 'ff66ff'
# 새로운 시트 > 동시에 타이틀
ws1 = wb.create_sheet('BK2_new')
ws2 = wb.create_sheet('BK3_new', 2)
ws2.sheet_properties.tabColor = 'dd0000'

# Dict 형태로 시트에 접근
new_ws = wb['BK_new']

# 모든 시트 정보 확인 ()
print(wb.sheetnames)

# 시트 복사
new_ws['A1'] = 'TEST'
target = wb.copy_worksheet(new_ws)
target.title = 'Copied sheet'


wb.save('C:/Users/PC/Documents/simbyungki/git/python_study/rpa/excel/sample.xlsx')
