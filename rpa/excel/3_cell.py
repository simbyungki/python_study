from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = 'BKSheet'
ws['A1'] = 1
ws['A2'] = 2
ws['A3'] = 3

ws['B1'] = 4
ws['B2'] = 5
ws['B3'] = 6

# A1 셀의 정보를 출력
print(ws['A1'])

# A1 셀의 값을 출력
# print(ws['A1'].value)
# 셀 값이 없는 경우 None
# print(ws['A10'].value)

# row = 1, 2, 3 ...
# column = 1(A), 2(B), 3(C) ...
# print(ws.cell(row=1, column=2).value)

# ws['C1'].value = 10
c = ws.cell(column=3, row=1, value=10)
# ws['C1']
# print(c.value)


from random import *
index = 1
for x in range(1, 11) :
	for y in range(1, 11) :
		# ws.cell(row=x, column=y, value=randint(0, 100))
		ws.cell(row=x, column=y, value=index)
		index += 1


wb.save('C:/Users/PC/Documents/simbyungki/git/python_study/rpa/excel/sample.xlsx')