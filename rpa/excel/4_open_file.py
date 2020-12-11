from openpyxl import load_workbook
# 파일 로드
wb = load_workbook('C:/Users/PC/Documents/simbyungki/git/python_study/rpa/excel/sample.xlsx')
# 활성화된 시트
ws = wb.active

# 출력
# for x in range(1, 11) : 
# 	for y in range(1, 11) : 
# 		print(ws.cell(row=x, column=y).value, end=' ')
# 	print()

# cell 갯수를 모를 때
for x in range(1, ws.max_row + 1) : 
	for y in range(1, ws.max_column + 1) : 
		print(ws.cell(row=x, column=y).value, end=' ')
	print()